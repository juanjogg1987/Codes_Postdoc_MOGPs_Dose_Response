#import xlwt
# import xlrd
# from xlutils.copy import copy
#
# # load the excel file
# rb = xlrd.open_workbook('UserBook.xls')
#
# # copy the contents of excel file
# wb = copy(rb)
#
# # open the first sheet
# w_sheet = wb.get_sheet(0)
#
# # row number = 0 , column number = 1
# w_sheet.write(5, 1, 'Modified2!')
#
# # save the file
# wb.save('UserBook.xls')


import openpyxl
def main():
    book = openpyxl.Workbook()
    book.create_sheet('Sample')# Acquire a sheet by its name
    sheet = book.get_sheet_by_name('Sample')# Writing to sheet
    sheet.cell(row=1, column=1).value = 'sample'
    book.save('Sample.xlsx')

    if __name__ == '__main__':
        main()